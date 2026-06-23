"""Export Service - Export reports to Excel/PDF"""
import io
import re
from datetime import date, datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
import calendar
from app.models.employee import Employee
from app.models.attendance import Attendance
from app.models.time_off import LeaveRequest


class ExportService:
    """Service for exporting reports to various formats"""
    
    @staticmethod
    def _format_working_hours(hours: float) -> str:
        """Format working hours: show minutes if < 1 hour, else show hours"""
        if hours is None or hours == 0:
            return "0 phút"
        
        if hours < 1:
            minutes = int(hours * 60)
            return f"{minutes} phút"
        else:
            return f"{hours:.2f}h"
    
    @staticmethod
    def _sanitize_sheet_name(name: str, max_length: int = 31) -> str:
        """
        Sanitize sheet name to be valid for Excel
        
        Excel sheet name rules:
        - Cannot contain: / \ ? * [ ] :
        - Maximum 31 characters
        - Cannot be empty
        """
        # Replace invalid characters with dash
        invalid_chars = r'[/\\?*\[\]:]'
        sanitized = re.sub(invalid_chars, '-', name)
        
        # Trim to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        
        # Ensure not empty
        if not sanitized.strip():
            sanitized = "Sheet1"
        
        return sanitized
    
    @staticmethod
    def export_to_excel(report_data: dict, report_type: str):
        """Export report data to Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set title (sanitize sheet name to remove invalid characters)
        if report_type == 'daily':
            # Format date for display: convert YYYY-MM-DD to DD/MM/YYYY
            date_iso = report_data.get('date', '')
            date_display = date_iso
            if date_iso:
                try:
                    date_obj = datetime.strptime(date_iso, '%Y-%m-%d').date()
                    date_display = date_obj.strftime('%d/%m/%Y')
                except:
                    pass
            
            # Sheet title: keep ISO format (YYYY-MM-DD) which doesn't have invalid chars
            sheet_title = f"Báo cáo ngày {date_iso}"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            title = f"BÁO CÁO CHẤM CÔNG NGÀY {date_display}"
        elif report_type == 'weekly':
            sheet_title = "Báo cáo tuần"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            start_date = report_data.get('start_date', '')
            end_date = report_data.get('end_date', '')
            title = f"BÁO CÁO CHẤM CÔNG TUẦN {start_date} - {end_date}"
        else:
            month = report_data.get('month', '')
            year = report_data.get('year', '')
            try:
                month = int(month)
                year = int(year)
            except Exception:
                month = month
                year = year
            # Sheet title: use dash instead of slash
            sheet_title = f"Báo cáo tháng {month}-{year}"
            ws.title = ExportService._sanitize_sheet_name(sheet_title)
            # Title in cell: can use slash for display
            title = f"BÁO CÁO CHẤM CÔNG THÁNG {month}/{year}"
        
        # Title style
        title_fill = PatternFill(start_color="667eea", end_color="764ba2", fill_type="solid")
        title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        
        # Header style
        header_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write title (merge adjusted for current columns)
        ws.merge_cells('A1:J1')
        ws['A1'] = title
        ws['A1'].fill = title_fill
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Write summary
        row = 3
        ws[f'A{row}'] = 'TỔNG HỢP'
        ws[f'A{row}'].font = Font(name="Arial", size=12, bold=True)
        row += 1
        
        if report_type == 'daily':
            summary_data = [
                ['Tổng nhân viên', report_data.get('total_employees')],
                ['Đã chấm công', report_data.get('checked_in')],
                ['Có mặt', report_data.get('present')],
                ['Đi muộn', report_data.get('late')],
                ['Vắng mặt', report_data.get('absent')],
                ['Tỷ lệ chấm công', f"{report_data.get('attendance_rate', '')}%"],
                ['Tổng công hành chính', f"{ExportService._format_working_hours(report_data.get('total_working_hours', 0)).rstrip('h')} công"],
                ['Tổng giờ tăng ca (thực tế)', f"{ExportService._format_working_hours(report_data.get('total_overtime_actual_hours', 0)).rstrip('h')} giờ"],
                ['Tổng công (tổng) = Hành chính + OT', f"{ExportService._format_working_hours(report_data.get('total_working_hours', 0) + report_data.get('total_overtime_hours', 0)).rstrip('h')} công"],
            ]
        elif report_type == 'weekly':
            summary_data = [
                ['Tổng có mặt', report_data['week_total_present']],
                ['Tổng đi muộn', report_data['week_total_late']],
                ['Tổng vắng mặt', report_data['week_total_absent']],
                ['Tổng công', f"{report_data['week_total_hours']:.2f} công"],
                ['Tỷ lệ trung bình', f"{report_data['average_attendance_rate']}%"],
            ]
        else:
            summary_data = [
                ['Tổng có mặt', report_data['total_present']],
                ['Tổng đi muộn', report_data['total_late']],
                ['Tổng vắng mặt', report_data['total_absent']],
                ['Tổng công', f"{ExportService._format_working_hours(report_data['total_working_hours']).rstrip('h')} công"],
                ['Tỷ lệ chấm công', f"{report_data['average_attendance_rate']}%"],
                ['Tỷ lệ đúng giờ', f"{report_data['on_time_rate']}%"],
            ]
        
        for item in summary_data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            row += 1
        
        row += 1
        
        # Write table header
        if report_type == 'daily':
            # Removed 'OT (giờ)' column per user's request
            headers = ['Nhân viên', 'Mã NV', 'Check-in', 'Check-out', 'Giờ làm', 'OT start', 'OT end', 'OT (công)', 'Tổng công', 'Trạng thái']
            cols = ['A','B','C','D','E','F','G','H','I','J']
            for idx, h in enumerate(headers):
                ws[f'{cols[idx]}{row}'] = h
        elif report_type == 'weekly':
            headers = ['Ngày', 'Có mặt', 'Đi muộn', 'Vắng mặt', 'Giờ làm']
            ws[f'A{row}'] = headers[0]
            ws[f'B{row}'] = headers[1]
            ws[f'C{row}'] = headers[2]
            ws[f'D{row}'] = headers[3]
            ws[f'E{row}'] = headers[4]
        else:
            # Monthly matrix format
            try:
                days_in_month = calendar.monthrange(int(year), int(month))[1]
            except Exception:
                days_in_month = 31
            headers = ['STT', 'Nhân viên', 'Chức vụ'] + [f"{d:02d}" for d in range(1, days_in_month + 1)] + ['Tổng công', 'Tổng OT', 'Đi muộn', 'Về sớm', 'Nghỉ phép']
            for idx, h in enumerate(headers):
                ws[f'{get_column_letter(idx+1)}{row}'] = h
        
        # Apply header style
        for col in range(1, len(headers) + 1):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Write data
        if report_type == 'daily':
            for att in report_data.get('attendances', [])[:100]:  # Limit to 100 rows
                ws[f'A{row}'] = att.get('employee_name', '')
                ws[f'B{row}'] = att.get('employee_code', '')
                # check-in
                check_in = att.get('check_in_time', '')
                if check_in:
                    try:
                        dt = datetime.fromisoformat(check_in.replace('Z', '+00:00'))
                        ws[f'C{row}'] = dt.strftime('%H:%M:%S')
                    except:
                        ws[f'C{row}'] = (check_in[:5] if len(check_in) > 5 else check_in)
                else:
                    ws[f'C{row}'] = 'N/A'

                # check-out
                check_out = att.get('check_out_time', '')
                if check_out:
                    try:
                        dt = datetime.fromisoformat(check_out.replace('Z', '+00:00'))
                        ws[f'D{row}'] = dt.strftime('%H:%M:%S')
                    except:
                        ws[f'D{row}'] = (check_out[:5] if len(check_out) > 5 else check_out)
                else:
                    ws[f'D{row}'] = 'N/A'

                # working hours
                ws[f'E{row}'] = f"{att.get('working_hours', 0)}h"

                # OT start / end / hours / points
                ot_start = att.get('overtime_check_in_time')
                if ot_start:
                    try:
                        dt = datetime.fromisoformat(ot_start.replace('Z', '+00:00'))
                        ws[f'F{row}'] = dt.strftime('%H:%M:%S')
                    except:
                        ws[f'F{row}'] = (ot_start[:5] if len(ot_start) > 5 else ot_start)
                else:
                    ws[f'F{row}'] = 'N/A'

                ot_end = att.get('overtime_check_out_time')
                if ot_end:
                    try:
                        dt = datetime.fromisoformat(ot_end.replace('Z', '+00:00'))
                        ws[f'G{row}'] = dt.strftime('%H:%M:%S')
                    except:
                        ws[f'G{row}'] = (ot_end[:5] if len(ot_end) > 5 else ot_end)
                else:
                    ws[f'G{row}'] = 'N/A'

                # overtime_hours (OT points) — OT actual hours column removed
                ot_points = att.get('overtime_hours', 0) or 0
                ws[f'H{row}'] = f"{ot_points:.2f}"

                # total work = admin_hours (hours) + OT points (công)
                total_display = (att.get('working_hours', 0) or 0) + ot_points
                ws[f'I{row}'] = f"{total_display:.2f}"

                ws[f'J{row}'] = att.get('status_label') or {
                    'present': 'Đi làm đúng giờ',
                    'late': 'Đi làm muộn',
                    'early_leave': 'Nghỉ sớm',
                    'absent': 'Vắng mặt',
                    'missing_check_in': 'Thiếu check-in',
                    'missing_check_out': 'Thiếu check-out',
                }.get(att.get('status', ''), att.get('status', ''))

                # Apply border
                for col in range(1, len(headers) + 1):
                    ws[f'{get_column_letter(col)}{row}'].border = border

                row += 1
        elif report_type == 'weekly':
            for day in report_data.get('daily_stats', []):
                ws[f'A{row}'] = day.get('date', '')
                ws[f'B{row}'] = day.get('present', 0)
                ws[f'C{row}'] = day.get('late', 0)
                ws[f'D{row}'] = day.get('absent', 0)
                ws[f'E{row}'] = f"{day.get('working_hours', 0)}h"
                for col in range(1, len(headers) + 1):
                    ws[f'{get_column_letter(col)}{row}'].border = border
                row += 1
        else:
            # Monthly matrix: build per-employee per-day values
            try:
                days_in_month = calendar.monthrange(int(year), int(month))[1]
            except Exception:
                days_in_month = 31

            employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
            idx_no = 1
            for emp in employees:
                col_idx = 1
                ws[f'{get_column_letter(col_idx)}{row}'] = idx_no
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = emp.name
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = getattr(emp, 'title', '') or getattr(emp, 'position', '') or ''
                col_idx += 1

                total_cong = 0.0
                total_ot = 0.0
                late_count = 0
                early_count = 0
                leave_count = 0

                for d in range(1, days_in_month + 1):
                    cur_date = date(int(year), int(month), d)
                    att = Attendance.query.filter_by(employee_id=emp.id, date=cur_date).first()
                    cell_val = ''
                    if att:
                        try:
                            att.calculate_working_hours()
                        except Exception:
                            pass
                        wh = att.working_hours or 0
                        if wh >= 8:
                            cell_val = 1
                            total_cong += 1
                        elif wh > 0:
                            cell_val = 0.5
                            total_cong += 0.5
                        else:
                            cell_val = 0

                        try:
                            od = att.to_dict().get('overtime_actual_hours') or 0
                        except Exception:
                            od = 0
                        total_ot += od

                        if att.status == 'late':
                            late_count += 1
                        if att.status == 'early_leave':
                            early_count += 1
                    else:
                        lr = LeaveRequest.query.filter(LeaveRequest.employee_id == emp.id, LeaveRequest.status == 'approved', LeaveRequest.start_date <= cur_date, LeaveRequest.end_date >= cur_date).first()
                        if lr:
                            cell_val = 'P'
                            leave_count += 1
                        else:
                            cell_val = ''

                    ws[f'{get_column_letter(col_idx)}{row}'] = cell_val
                    col_idx += 1

                # Totals
                ws[f'{get_column_letter(col_idx)}{row}'] = total_cong
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = total_ot
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = late_count
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = early_count
                col_idx += 1
                ws[f'{get_column_letter(col_idx)}{row}'] = leave_count

                for c in range(1, len(headers) + 1):
                    ws[f'{get_column_letter(c)}{row}'].border = border

                row += 1
                idx_no += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_num in range(1, row):
                cell_value = ws[f'{column_letter}{row_num}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Add a guidance sheet with usage instructions
        try:
            guide = wb.create_sheet(title=ExportService._sanitize_sheet_name('Hướng dẫn'))
            guide['A1'] = 'Hướng dẫn sử dụng file báo cáo'
            guide['A2'] = '1) Cột STT: số thứ tự'
            guide['A3'] = '2) Cột 01..NN: điểm công cho từng ngày (1 = công, 0.5 = nửa ngày, P = nghỉ phép)'
            guide['A4'] = '3) Cột Tổng công: tổng điểm công trong tháng (1 = đủ ngày)'
            guide['A5'] = '4) Cột Tổng OT: tổng giờ OT thực tế (giờ)'
            guide['A6'] = '5) Sử dụng Excel để lọc, pivot, hoặc nhân 1 công = 8 giờ khi cần chuyển đổi sang giờ.'
        except Exception:
            pass
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

